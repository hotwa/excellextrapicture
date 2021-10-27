#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@file        :xlsx_pic.py
@Description:       :extract excell file picture files
@Date     :2021/01/16 10:55:18
@Author      :hotwa
@version      :1.0
'''

import base64
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader # picture dependency
from openpyxl.utils import get_column_letter, column_index_from_string # number transform letter
from io import BytesIO



class xlsx_pic(object):
    __slots__ = [
        'openpyxl_wb','all_sheets','filename','lstd'
    ]

    def __init__(
        self,
        filename: dict(type=str, help="XLSX file name or path")):
        self.filename = filename
        self.openpyxl_wb = load_workbook(filename)
        self.all_sheets = self.openpyxl_wb.sheetnames
        self.lstd = {j:i for i,j in enumerate(self.all_sheets)}

    @property
    def sheetname(self):
        return self.all_sheets[self.sheetnum]

    @sheetname.setter
    def sheetname(self,value:str):
        if isinstance(value,str): self.openpyxl_wb.active = self.openpyxl_wb[value]

    @property
    def sheetnum(self):
        return self.lstd[self.openpyxl_wb.active.title]

    @sheetnum.setter
    def sheetnum(self,value:int):
        if isinstance(value,int): self.openpyxl_wb.active = value

    @staticmethod
    def get_cell_pic(filename,sheetnum,position,new_name=None,base64c=None)->bytes:
        """get_cell_pic [summary]
        
        [extended_summary]
        
        :param filename: xlsx filename
        :type filename: [str]
        :param sheetnum: sequence of sheet names
        :type sheetnum: [int]
        :param position: picture cell position, like E2
        :type position: [str]
        :param new_name: [save file name], defaults to None, like 'savepic.png'
        :type new_name: [str], optional
        :param base64c: [transfer to base64], defaults to None
        :type base64c: [bool], optional
        :return: [base64 code or None]
        :rtype: [str]
        """
        wb = load_workbook(filename)
        ws = wb.active # 调用得到正在运行的工作表
        ws.title # 调用当前运行的工作的名称
        sheet = wb[wb.sheetnames[sheetnum]] # read first sheet
        image_loader = SheetImageLoader(sheet)
        if image_loader.image_in(position):
            image = image_loader.get(position)
            if new_name:
                # 储存文件
                image.save(new_name)
            else:
                # 返回二进制内容
                bytesIO = BytesIO()
                image.save(bytesIO, format='PNG')
                if base64c:
                    return base64.b64encode(bytesIO.getvalue()).decode() # default to string
                    # return '1' # for test
                else:
                    return bytesIO.getvalue()
        else:
            return None

    def read_cell(self,pos:dict(type=str,help='excell position like E2')):
        """read_cell read_cell openpyxl methods get cell values
        
        [picture cell return bytes(format:PNG)]
        
        :param pos: [cell description], defaults to str,help='excell position like E2')
        :type pos: [string], optional
        """
        sheet = self.openpyxl_wb[self.all_sheets[self.sheetnum]]
        image_loader = SheetImageLoader(sheet)
        if image_loader.image_in(pos): # picture cell return bytes
            bytes_value = self.get_cell_pic(filename=self.filename,sheetnum=self.sheetnum,position=pos,base64c=True)
            return bytes_value
        else:
            return sheet[pos].value

    def read_row(self,row_num:int,base64c=None,read_cell_picture: bool=False,pic_column:list=[]):
        """read_row 相比read_sheet_line方法，空cell返回None，默认数字类型自动转化为int
        
        [extended_summary]
        
        :param row_num: [行数]
        :type row_num: [int]
        :param read_cell_picture: [自动尝试读取单元格图片], defaults to False
        :type read_cell_picture: [bool], optional
        :param pic_column: [read columns of string list], defaults to []
        :type pic_column: [list], optional
        :return: [返回改行读取的列表]
        :rtype: [list]
        """
        sheet = self.openpyxl_wb.active
        row = sheet[row_num]
        row_num_sum = len(row)
        _row_list = []
        for i in range(row_num_sum):
            _row_list.append(row[i].value)
        _ll = [False if i == None else 1 for i in _row_list]
        if sum(_ll) == 0:
            return None # 此行为空行
        else:
            if read_cell_picture:
                if pic_column == []:
                    row_list = self.__cell_pic_try(row_num=row_num,current_row_list=_row_list,base64c=base64c)
                else:
                    row_list = self.__cell_pic_try(row_num=row_num,current_row_list=_row_list,pic_column=pic_column,base64c=base64c)
                return row_list
            return _row_list

    def __cell_pic_try(self,row_num,current_row_list,base64c=None,pic_column=[])-> list:
        """__cell_pic_try 尝试对改行的None进行图片提取返回二进制
        
        [extended_summary]
        
        :param row_num: [行号]
        :type row_num: [int]
        """
        # print(f'当前行读取内容{current_row_list},{row_num}')
        if None in current_row_list:
            while (None in current_row_list):
                _pos = current_row_list.index(None)
                letter = get_column_letter(_pos+1)
                pos=letter+str(row_num)
                # print(f'尝试定位{pos}位置的图片')
                if pic_column == []:
                    res = self.get_cell_pic(filename=self.filename,sheetnum=self.sheetnum,position=pos,base64c = base64c)
                    if res:
                        # print(f'定位图片成功{pos}')
                        current_row_list[_pos] = res
                    else:
                        current_row_list[_pos] = False
                elif letter in pic_column:
                    res = self.get_cell_pic(filename=self.filename,sheetnum=self.sheetnum,position=pos,base64c = base64c)
                    if res:
                        # print(f'定位图片成功{pos}')
                        current_row_list[_pos] = res
                    else:
                        current_row_list[_pos] = False
                else:
                    # print(f'跳过{pos}位置的图片')
                    current_row_list[_pos] = False
            return current_row_list
        else:
            return None

# for test 
if __name__ == '__main__':
    x = xlsx_pic('./test.xlsx')